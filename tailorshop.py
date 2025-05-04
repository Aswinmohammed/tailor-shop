import openpyxl
from openpyxl import Workbook
import os
from tkinter import messagebox
from tkinter import Menu, Toplevel, END
from customtkinter import *
from PIL import Image
from datetime import datetime

class TailorShopApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JR FASHION MART")
        self.root.geometry("400x520")
        self.excel_file = r"D:\Tailor_Shop\tailor_shop_orders.xlsx"
        self.orders = {}
        self.load_orders()

        set_appearance_mode("light")  # Set the appearance mode

        # Menu bar
        menu_bar = Menu(root)
        root.config(menu=menu_bar)

        main_menu = Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Main Window", menu=main_menu)
        main_menu.add_command(label="Main Window", command=self.show_main_window)

        expense_income_menu = Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Expenses & Income", menu=expense_income_menu)
        expense_income_menu.add_command(label="Expenses & Income", command=self.show_expense_income_window)

        # Correctly specify the path to your background image
        background_image_path = r"D:\Tailor_Shop\img.jpeg"

        # Load and set the background image using CTkImage
        self.background_image = CTkImage(dark_image=Image.open(background_image_path), size=(400, 520))
        self.bg_label = CTkLabel(root, image=self.background_image)
        self.bg_label.place(relwidth=1, relheight=1)

        # Custom font and styles
        self.bold_font = ("Helvetica", 12, "bold")

        # Entry fields
        self.create_label(root, "Order No:", 0, 0)
        self.order_no_entry = self.create_entry(root, 0, 1)

        self.create_label(root, "Name:", 1, 0)
        self.name_entry = self.create_entry(root, 1, 1)

        self.create_label(root, "Phone Number:", 2, 0)
        self.phone_entry = self.create_entry(root, 2, 1)

        self.create_label(root, "Total Charge:", 3, 0)
        self.total_charge_entry = self.create_entry(root, 3, 1)

        self.create_label(root, "Advance:", 4, 0)
        self.advance_entry = self.create_entry(root, 4, 1)

        self.create_label(root, "Balance:", 5, 0)
        self.balance_entry = self.create_entry(root, 5, 1)
        self.balance_entry.configure(state='normal')

        self.create_label(root, "Date:", 6, 0)
        self.date_entry = self.create_entry(root, 6, 1)
        self.date_entry.insert(0, "DD-MM-YYYY")
        self.date_entry.bind("<FocusIn>", lambda event: self.clear_placeholder(event, "DD-MM-YYYY"))
        self.date_entry.bind("<FocusOut>", lambda event: self.add_placeholder(event, "DD-MM-YYYY"))

        self.create_label(root, "Delivery Date:", 7, 0)
        self.delivery_date_entry = self.create_entry(root, 7, 1)
        self.delivery_date_entry.insert(0, "DD-MM-YYYY")
        self.delivery_date_entry.bind("<FocusIn>", lambda event: self.clear_placeholder(event, "DD-MM-YYYY"))
        self.delivery_date_entry.bind("<FocusOut>", lambda event: self.add_placeholder(event, "DD-MM-YYYY"))

        # Buttons
        self.add_button = CTkButton(root, text="Add Order", command=self.add_order, fg_color="#50C878")
        self.add_button.place(x=120, y=325)

        self.search_button = CTkButton(root, text="Search", command=self.search_order, fg_color="#007FFF")
        self.search_button.place(x=120, y=365)

        self.clear_button = CTkButton(root, text="Clear", command=self.clear_entries, fg_color="#2F4F4F")
        self.clear_button.place(x=120, y=405)

        self.update_button = CTkButton(root, text="Update Balance", command=self.update_order, fg_color="#FF6347")
        self.update_button.place(x=120, y=445)

        # Bind the calculation of balance to changes in total charge and advance
        self.total_charge_entry.bind("<KeyRelease>", self.calculate_balance)
        self.advance_entry.bind("<KeyRelease>", self.calculate_balance)

    def create_label(self, parent, text, row, column):
        label = CTkLabel(parent, text=text, font=self.bold_font, bg_color="#B0E0E6")
        label.grid(row=row, column=column, padx=10, pady=5)

    def create_entry(self, parent, row, column):
        entry = CTkEntry(parent)
        entry.grid(row=row, column=column, padx=10, pady=5)
        return entry

    def clear_placeholder(self, event, placeholder):
        if event.widget.get() == placeholder:
            event.widget.delete(0, END)
            event.widget.config(fg='grey')

    def add_placeholder(self, event, placeholder):
        if event.widget.get() == "":
            event.widget.insert(0, placeholder)
            event.widget.config(fg='grey')

    def load_orders(self):
        try:
            if os.path.exists(self.excel_file):
                workbook = openpyxl.load_workbook(self.excel_file)
                if "Orders" in workbook.sheetnames:
                    order_sheet = workbook["Orders"]
                    for row in order_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) == 8:
                            order_no, name, phone, total_charge, advance, balance, date, delivery_date = row
                            if isinstance(name, str):
                                self.orders[name.lower()] = {
                                    "order_no": order_no,
                                    "name": name,
                                    "phone": phone,
                                    "total_charge": total_charge,
                                    "advance": advance,
                                    "balance": balance,
                                    "date": date,
                                    "delivery_date": delivery_date
                                }
                else:
                    workbook.create_sheet("Orders")
                    order_sheet = workbook["Orders"]
                    order_sheet.append(["Order No", "Name", "Phone Number", "Total Charge", "Advance", "Balance", "Date", "Delivery Date"])
                    workbook.save(self.excel_file)

                if "ExpensesIncome" not in workbook.sheetnames:
                    workbook.create_sheet("ExpensesIncome")
                    expense_income_sheet = workbook["ExpensesIncome"]
                    expense_income_sheet.append(["Expense", "Income", "Date"])
                    workbook.save(self.excel_file)
            else:
                workbook = Workbook()
                order_sheet = workbook.create_sheet("Orders")
                order_sheet.append(["Order No", "Name", "Phone Number", "Total Charge", "Advance", "Balance", "Date", "Delivery Date"])
                expense_income_sheet = workbook.create_sheet("ExpensesIncome")
                expense_income_sheet.append(["Expense", "Income", "Date"])
                workbook.save(self.excel_file)
        except PermissionError:
            messagebox.showerror("Permission Error", f"Permission denied for file: {self.excel_file}")

    def add_order(self):
        order_no = self.order_no_entry.get()
        name = self.name_entry.get()
        phone = self.phone_entry.get()
        total_charge = self.total_charge_entry.get()
        advance = self.advance_entry.get()
        balance = self.balance_entry.get()
        date = self.date_entry.get()
        delivery_date = self.delivery_date_entry.get()

        if not order_no or not name or not phone or not total_charge or not advance or not balance or not date or not delivery_date:
            messagebox.showerror("Input Error", "All fields are required")
            return

        if name.lower() in self.orders:
            messagebox.showerror("Duplicate Error", "Order for this name already exists")
            return

        self.orders[name.lower()] = {
            "order_no": order_no,
            "name": name,
            "phone": phone,
            "total_charge": total_charge,
            "advance": advance,
            "balance": balance,
            "date": date,
            "delivery_date": delivery_date
        }

        self.save_order_to_excel(order_no, name, phone, total_charge, advance, balance, date, delivery_date)
        messagebox.showinfo("Success", "Order added successfully")
        self.clear_entries()

    def search_order(self):
        name = self.name_entry.get().lower()
        if name in self.orders:
            order = self.orders[name]
            self.order_no_entry.delete(0, END)
            self.order_no_entry.insert(0, order["order_no"])

            self.phone_entry.delete(0, END)
            self.phone_entry.insert(0, order["phone"])

            self.total_charge_entry.delete(0, END)
            self.total_charge_entry.insert(0, order["total_charge"])

            self.advance_entry.delete(0, END)
            self.advance_entry.insert(0, order["advance"])

            self.balance_entry.configure(state='normal')
            self.balance_entry.delete(0, END)
            self.balance_entry.insert(0, order["balance"])

            self.date_entry.delete(0, END)
            self.date_entry.insert(0, order["date"])

            self.delivery_date_entry.delete(0, END)
            self.delivery_date_entry.insert(0, order["delivery_date"])
        else:
            messagebox.showerror("Not Found", "Order not found")

    def calculate_balance(self, event=None):
        try:
            total_charge = float(self.total_charge_entry.get())
            advance = float(self.advance_entry.get())
            balance = total_charge - advance
            self.balance_entry.configure(state='normal')
            self.balance_entry.delete(0, END)
            self.balance_entry.insert(0, f"{balance:.2f}")
            self.balance_entry.configure(state='disabled')
        except ValueError:
            pass

    def save_order_to_excel(self, order_no, name, phone, total_charge, advance, balance, date, delivery_date):
        workbook = openpyxl.load_workbook(self.excel_file)
        sheet = workbook["Orders"]
        sheet.append([order_no, name, phone, total_charge, advance, balance, date, delivery_date])
        workbook.save(self.excel_file)

    def update_order(self):
        name = self.name_entry.get().lower()
        if name in self.orders:
            order = self.orders[name]
            order_no = order["order_no"]
            phone = order["phone"]
            total_charge = self.total_charge_entry.get()
            advance = self.advance_entry.get()
            balance = self.balance_entry.get()
            date = self.date_entry.get()
            delivery_date = self.delivery_date_entry.get()

            order["total_charge"] = total_charge
            order["advance"] = advance
            order["balance"] = balance
            order["date"] = date
            order["delivery_date"] = delivery_date

            self.orders[name] = order

            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook["Orders"]
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == order["name"]:
                    row[0].value = order_no
                    row[1].value = order["name"]
                    row[2].value = phone
                    row[3].value = total_charge
                    row[4].value = advance
                    row[5].value = balance
                    row[6].value = date
                    row[7].value = delivery_date
                    break

            workbook.save(self.excel_file)
            messagebox.showinfo("Success", "Order updated successfully")
            self.clear_entries()
        else:
            messagebox.showerror("Not Found", "Order not found")

    def clear_entries(self):
        self.order_no_entry.delete(0, END)
        self.name_entry.delete(0, END)
        self.phone_entry.delete(0, END)
        self.total_charge_entry.delete(0, END)
        self.advance_entry.delete(0, END)
        self.balance_entry.configure(state='normal')
        self.balance_entry.delete(0, END)
        self.date_entry.delete(0, END)
        self.date_entry.insert(0, "DD-MM-YYYY")
        self.delivery_date_entry.delete(0, END)
        self.delivery_date_entry.insert(0, "DD-MM-YYYY")

    def show_main_window(self):
        self.root.deiconify()

    def show_expense_income_window(self):
        expense_income_window = Toplevel(self.root)
        expense_income_window.title("Expenses & Income")
        expense_income_window.geometry("400x400")

        # Labels for expenses and income
        expense_label = CTkLabel(expense_income_window, text="Expenses", font=self.bold_font)
        expense_label.pack(pady=10)

        income_label = CTkLabel(expense_income_window, text="Income", font=self.bold_font)
        income_label.pack(pady=10)

        # Entry fields for expenses and income
        expense_entry = CTkEntry(expense_income_window)
        expense_entry.pack(pady=5)

        income_entry = CTkEntry(expense_income_window)
        income_entry.pack(pady=5)

        # Date labels and entry fields
        date_label = CTkLabel(expense_income_window, text="Date (DD-MM-YYYY)", font=self.bold_font)
        date_label.pack(pady=10)

        date_entry = CTkEntry(expense_income_window)
        date_entry.pack(pady=5)
        date_entry.insert(0, "DD-MM-YYYY")
        date_entry.bind("<FocusIn>", lambda event: self.clear_placeholder(event, "DD-MM-YYYY"))
        date_entry.bind("<FocusOut>", lambda event: self.add_placeholder(event, "DD-MM-YYYY"))

        # Save button
        save_button = CTkButton(expense_income_window, text="Save", command=lambda: self.save_expense_income(expense_entry.get(), income_entry.get(), date_entry.get()))
        save_button.pack(pady=20)

    def save_expense_income(self, expense, income, date):
        if not expense or not income or not date:
            messagebox.showerror("Input Error", "All fields are required")
            return

        messagebox.showinfo("Success", "Expenses and Income saved successfully")

    def show_expense_income_window(self):
        self.root.withdraw()
        expense_income_window = Toplevel(self.root)
        expense_income_window.title("Expenses & Income")
        expense_income_window.geometry("400x500")
        expense_income_window.resizable(False, False)

        # Load and set the background image using CTkImage
        bg_label = CTkLabel(expense_income_window, image=self.background_image)
        bg_label.place(relwidth=1, relheight=1)

        # Create a label for displaying the date
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        date_label = CTkLabel(expense_income_window, text=f"Date: {current_date}", font=self.bold_font, bg_color="#F67280")
        date_label.place(x=120, y=460)

        self.create_label(expense_income_window, "Expense Description:", 0, 0)
        self.expense_description_entry = self.create_entry(expense_income_window, 0, 1)

        self.create_label(expense_income_window, "Expense Amount:", 1, 0)
        self.expense_amount_entry = self.create_entry(expense_income_window, 1, 1)

        self.create_label(expense_income_window, "Income Description:", 2, 0)
        self.income_description_entry = self.create_entry(expense_income_window, 2, 1)

        self.create_label(expense_income_window, "Income Amount:", 3, 0)
        self.income_amount_entry = self.create_entry(expense_income_window, 3, 1)

        self.save_expense_button = CTkButton(expense_income_window, text="Save Expense", command=self.save_expense, fg_color="#00A86B")
        self.save_expense_button.place(x=30, y=180)

        self.save_income_button = CTkButton(expense_income_window, text="Save Income", command=self.save_income, fg_color="#00A86B")
        self.save_income_button.place(x=230, y=180)

        self.total_income_label = CTkLabel(expense_income_window, text="Total Income: 0", font=self.bold_font, bg_color="#B0E0E6")
        self.total_income_label.place(x=140, y=230)

        self.total_expense_label = CTkLabel(expense_income_window, text="Total Expenses: 0", font=self.bold_font, bg_color="#B0E0E6")
        self.total_expense_label.place(x=140, y=280)

        self.show_totals_button = CTkButton(expense_income_window, text="Show Totals", command=self.show_totals, fg_color="#007FFF")
        self.show_totals_button.place(x=120, y=330)

        self.clear_button = CTkButton(expense_income_window, text="Clear All", command=self.clear_all_expenses_incomes, fg_color="maroon")
        self.clear_button.place(x=120, y=370)

        self.back_button = CTkButton(expense_income_window, text="Back", command=lambda: self.back_to_main(expense_income_window), fg_color="navyblue")
        self.back_button.place(x=120, y=410)

    def save_expense(self):
        description = self.expense_description_entry.get()
        amount = self.expense_amount_entry.get()
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not description or not amount:
            messagebox.showerror("Input Error", "Both fields are required")
            return

        try:
            amount = float(amount)

        except ValueError:
            messagebox.showerror("Input Error", "Amount must be a number")
            return

        workbook = openpyxl.load_workbook(self.excel_file)

        if "Expenses & Income" not in workbook.sheetnames:
            workbook.create_sheet("Expenses & Income")
        sheet = workbook["Expenses & Income"]

        if sheet.max_row == 1:
            sheet.append(["Description", "Amount", "Type", "Date"])

        sheet.append([description, amount, "Expense", current_date])
        workbook.save(self.excel_file)
        messagebox.showinfo("Success", "Expense saved successfully")
        self.expense_description_entry.delete(0, END)
        self.expense_amount_entry.delete(0, END)

    def save_income(self):
        description = self.income_description_entry.get()
        amount = self.income_amount_entry.get()
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not description or not amount:
            messagebox.showerror("Input Error", "Both fields are required")
            return

        try:
            amount = float(amount)

        except ValueError:
            messagebox.showerror("Input Error", "Amount must be a number")
            return

        workbook = openpyxl.load_workbook(self.excel_file)

        if "Expenses & Income" not in workbook.sheetnames:
            workbook.create_sheet("Expenses & Income")
        sheet = workbook["Expenses & Income"]

        if sheet.max_row == 1:
            sheet.append(["Description", "Amount", "Type", "Date"])

        sheet.append([description, amount, "Income", current_date])
        workbook.save(self.excel_file)
        messagebox.showinfo("Success", "Income saved successfully")
        self.income_description_entry.delete(0, END)
        self.income_amount_entry.delete(0, END)

    def show_totals(self):
        workbook = openpyxl.load_workbook(self.excel_file, read_only=True)
        sheet = workbook["Expenses & Income"]

        total_expenses = 0
        total_income = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == "Expense":
                total_expenses += row[1]
            elif row[2] == "Income":
                total_income += row[1]

        self.total_expense_label.configure(text=f"Total Expenses: {total_expenses}")
        self.total_income_label.configure(text=f"Total Income: {total_income}")

    def clear_all_expenses_incomes(self):
        workbook = openpyxl.load_workbook(self.excel_file)
        sheet = workbook["Expenses & Income"]
        max_row = sheet.max_row

        for _ in range(2, max_row + 1):
            sheet.delete_rows(2)

        workbook.save(self.excel_file)
        self.total_expense_label.configure(text="Total Expenses: 0")
        self.total_income_label.configure(text="Total Income: 0")
        messagebox.showinfo("Cleared", "All expenses and incomes cleared successfully")

    def back_to_main(self, window):
        window.destroy()
        self.root.deiconify()

if __name__ == "__main__":
    root = CTk()
    app = TailorShopApp(root)
    root.resizable(False, False)
    root.mainloop()








